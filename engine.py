"""
项目管理核心引擎 v3
- 解析 Excel 进度表（保留依赖关系）
- 构建任务依赖图 (DAG)
- WORKDAY 计算逻辑 (排除周末)
- 用户输入开始日期 → 自动计算预计完成日期 + 级联更新后继
- 实际完成日期驱动状态：有值→COMPLETED，空→PLANNING，可手动→IN PROGRESS
- 数据持久化：修改自动保存到 JSON，加载时恢复
"""

import datetime
import json
import os
from collections import defaultdict, deque
from dataclasses import dataclass, asdict
from typing import Optional

import openpyxl


# ═══════════════════════════════════════════════════════════════
# 数据模型
# ═══════════════════════════════════════════════════════════════

@dataclass
class Task:
    """单个任务的数据模型"""
    task_id: int
    task_name: str
    duration: int                      # 工期（工作日天数）
    pred_task_ids: list[int]           # 前置任务 ID 列表
    pred_types: list[str]              # 依赖类型列表 (FS / SS)
    lag_days: list[int]                # 滞后天数列表
    start: datetime.date | None        # 计划开始日期（用户输入）
    planned_finish: datetime.date | None  # 预计完成日期（自动计算）
    actual_finish: datetime.date | None   # 实际完成日期（用户输入）
    status: str                        # PLANNING / IN PROGRESS / COMPLETED
    excel_row: int                     # 在 Excel Schedule 表中的行号

    def __hash__(self):
        return self.task_id

    def derive_status(self) -> str:
        """
        根据规则自动推导状态：
        - 有实际完成日期 → COMPLETED
        - 手动设置为 IN PROGRESS → 保持 IN PROGRESS（仅当无 actual_finish 时）
        - 否则 → PLANNING
        """
        if self.actual_finish is not None:
            return 'COMPLETED'
        if self.status == 'IN PROGRESS':
            return 'IN PROGRESS'
        return 'PLANNING'


# ═══════════════════════════════════════════════════════════════
# WORKDAY 工具函数
# ═══════════════════════════════════════════════════════════════

def is_weekend(d: datetime.date) -> bool:
    """判断是否为周末（周六=5, 周日=6）"""
    return d.weekday() >= 5


def workday(start_date: datetime.date, days: int) -> datetime.date:
    """
    模拟 Excel WORKDAY 函数。
    从 start_date 出发，偏移 days 个工作日。
    days>0 向后数工作日，days<0 向前数，days=0 返回当天。
    """
    if days == 0:
        return start_date
    current = start_date
    remaining = abs(days)
    step = 1 if days > 0 else -1
    while remaining > 0:
        current += datetime.timedelta(days=step)
        if not is_weekend(current):
            remaining -= 1
    return current


def calc_planned_finish(start_date: datetime.date, duration_days: int) -> datetime.date:
    """
    根据开始日期和工期计算预计完成日期。
    公式: Finish = WORKDAY(Start, Duration-1)
    例如: Start=周一, Duration=3 → Finish=周三
    """
    if duration_days <= 0:
        return start_date
    return workday(start_date, duration_days - 1)


# ═══════════════════════════════════════════════════════════════
# Excel 解析器
# ═══════════════════════════════════════════════════════════════

class ScheduleParser:
    """解析 Excel 进度表，提取任务结构（清空日期，保留依赖关系）"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.project_start_date: datetime.date | None = None
        self.project_address: str = ""
        self.tasks: dict[int, Task] = {}
        self.tasks_by_row: dict[int, Task] = {}

    def parse(self):
        """解析 Excel 文件，所有日期清空为 None"""
        self._wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = self._wb['Schedule']

        self.project_address = str(ws['A1'].value or '')
        self.project_start_date = self._parse_date(ws['C2'].value)

        for row_idx in range(4, ws.max_row + 1):
            tid = ws.cell(row=row_idx, column=1).value
            if tid is None or not isinstance(tid, (int, float)):
                continue
            tid = int(tid)
            task_name = str(ws.cell(row=row_idx, column=2).value or '').strip()

            duration_raw = ws.cell(row=row_idx, column=3).value
            duration = int(duration_raw) if duration_raw is not None else 0

            # 解析前置任务
            pred_raw = ws.cell(row=row_idx, column=4).value
            pred_ids = self._parse_pred_ids(pred_raw)

            # 解析依赖类型
            pred_type_raw = ws.cell(row=row_idx, column=5).value
            pred_types = self._parse_pred_types(pred_type_raw, len(pred_ids))

            # 解析滞后天数
            lag_raw = ws.cell(row=row_idx, column=6).value
            lag_days = self._parse_lag_days(lag_raw, len(pred_ids))

            # ★ 关键变更：所有日期清空为 None
            # 原有 Start(列G) 和 Finish(列H) 的值不读取，全部置空
            # 原 Status(列I) 也重置

            task = Task(
                task_id=tid,
                task_name=task_name,
                duration=duration,
                pred_task_ids=pred_ids,
                pred_types=pred_types,
                lag_days=lag_days,
                start=None,              # ★ 清空
                planned_finish=None,     # ★ 清空
                actual_finish=None,      # ★ 新增，初始为空
                status='PLANNING',       # ★ 默认为 PLANNING
                excel_row=row_idx,
            )
            self.tasks[tid] = task
            self.tasks_by_row[row_idx] = task

        return self

    def _parse_date(self, value) -> datetime.date | None:
        if value is None:
            return None
        if isinstance(value, datetime.datetime):
            return value.date()
        if isinstance(value, datetime.date):
            return value
        return None

    def _parse_pred_ids(self, raw) -> list[int]:
        if raw is None:
            return []
        if isinstance(raw, (int, float)):
            return [int(raw)]
        if isinstance(raw, str):
            raw = raw.strip()
            if not raw:
                return []
            parts = raw.replace('，', ',').split(',')
            result = []
            for p in parts:
                p = p.strip()
                if p:
                    try:
                        result.append(int(p))
                    except ValueError:
                        pass
            return result
        return []

    def _parse_pred_types(self, raw, count: int) -> list[str]:
        defaults = ['FS'] * count
        if raw is None:
            return defaults
        if isinstance(raw, str):
            raw = raw.strip().upper()
            if not raw:
                return defaults
            parts = [p.strip() for p in raw.replace('，', ',').split(',')]
            while len(parts) < count:
                parts.append('FS')
            return parts[:count]
        return defaults

    def _parse_lag_days(self, raw, count: int) -> list[int]:
        defaults = [0] * count
        if raw is None:
            return defaults
        if isinstance(raw, (int, float)):
            return [int(raw)] * count
        if isinstance(raw, str):
            raw = raw.strip()
            if not raw:
                return defaults
            parts = [p.strip() for p in raw.replace('，', ',').split(',')]
            result = []
            for p in parts:
                try:
                    result.append(int(p))
                except ValueError:
                    result.append(0)
            while len(result) < count:
                result.append(0)
            return result[:count]
        return defaults


# ═══════════════════════════════════════════════════════════════
# 依赖图 (DAG)
# ═══════════════════════════════════════════════════════════════

class DependencyGraph:
    """
    任务依赖图（有向无环图）
    """

    def __init__(self, tasks: dict[int, Task]):
        self.tasks = tasks
        self.successors: dict[int, list[int]] = defaultdict(list)
        self.predecessors: dict[int, list[int]] = defaultdict(list)
        self._build()

    def _build(self):
        for tid, task in self.tasks.items():
            for pred_id in task.pred_task_ids:
                if pred_id in self.tasks:
                    self.successors[pred_id].append(tid)
                    self.predecessors[tid].append(pred_id)

    def detect_cycle(self) -> list[int] | None:
        """DFS 检测循环依赖，返回环中节点列表或 None"""
        WHITE, GRAY, BLACK = 0, 1, 2
        color = {tid: WHITE for tid in self.tasks}
        parent = {}

        def dfs(u):
            color[u] = GRAY
            for v in self.successors.get(u, []):
                if color[v] == GRAY:
                    cycle = [v, u]
                    while u != v:
                        u = parent.get(u)
                        if u is None:
                            break
                        cycle.append(u)
                    cycle.reverse()
                    return cycle
                if color[v] == WHITE:
                    parent[v] = u
                    result = dfs(v)
                    if result:
                        return result
            color[u] = BLACK
            return None

        for tid in self.tasks:
            if color[tid] == WHITE:
                result = dfs(tid)
                if result:
                    return result
        return None

    def get_descendants(self, task_id: int) -> set[int]:
        """BFS 获取所有后继节点"""
        visited = set()
        queue = deque([task_id])
        while queue:
            current = queue.popleft()
            for succ in self.successors.get(current, []):
                if succ not in visited:
                    visited.add(succ)
                    queue.append(succ)
        return visited

    def topological_from(self, task_ids: set[int]) -> list[int]:
        """对给定集合进行拓扑排序"""
        in_degree = defaultdict(int)
        subgraph_edges = defaultdict(list)

        for tid in task_ids:
            for pred_id in self.predecessors.get(tid, []):
                if pred_id in task_ids:
                    in_degree[tid] += 1
                    subgraph_edges[pred_id].append(tid)

        queue = deque([tid for tid in task_ids if in_degree[tid] == 0])
        result = []

        while queue:
            current = queue.popleft()
            result.append(current)
            for succ in subgraph_edges.get(current, []):
                in_degree[succ] -= 1
                if in_degree[succ] == 0:
                    queue.append(succ)

        return result


# ═══════════════════════════════════════════════════════════════
# 日期级联更新引擎 v2
# ═══════════════════════════════════════════════════════════════

class ScheduleEngine:
    """
    进度表计算引擎 v2：
    - 用户输入开始日期 → 自动计算预计完成日期
    - 如果任务有前置依赖且前置有日期，开始日期自动根据依赖推导
    - 级联更新所有后继任务的开始和预计完成日期
    """

    def __init__(self, tasks: dict[int, Task], graph: DependencyGraph,
                 project_start: datetime.date):
        self.tasks = tasks
        self.graph = graph
        self.project_start = project_start

    def compute_start_date(self, task: Task) -> datetime.date | None:
        """
        根据前置任务计算当前任务的开始日期。
        如果所有前置任务的 planned_finish/start 都为空，返回 None（无法计算）。
        
        FS: start = WORKDAY(pred.planned_finish, lag+1)
        SS: start = WORKDAY(pred.start, lag)
        """
        if not task.pred_task_ids:
            # 无前置任务：无法自动推导，必须由用户手动输入
            return None

        candidate_starts = []
        all_none = True

        for i, pred_id in enumerate(task.pred_task_ids):
            pred = self.tasks.get(pred_id)
            if pred is None:
                continue

            pred_type = task.pred_types[i] if i < len(task.pred_types) else 'FS'
            lag = task.lag_days[i] if i < len(task.lag_days) else 0

            if pred_type == 'FS':
                if pred.planned_finish is not None:
                    all_none = False
                    candidate_starts.append(workday(pred.planned_finish, lag + 1))
            elif pred_type == 'SS':
                if pred.start is not None:
                    all_none = False
                    candidate_starts.append(workday(pred.start, lag))
            else:
                if pred.planned_finish is not None:
                    all_none = False
                    candidate_starts.append(workday(pred.planned_finish, lag + 1))

        if all_none:
            return None
        return max(candidate_starts) if candidate_starts else None

    def update_task_start(self, task_id: int, new_start: datetime.date) -> dict:
        """
        ★ 核心方法：用户输入某个任务的开始日期。
        1. 设置该任务的 start
        2. 自动计算 planned_finish = WORKDAY(start, duration-1)
        3. 级联更新所有后继任务的 start 和 planned_finish
        
        返回: {
            'updated': {task_id: {old/new start/planned_finish}},
            'error': str | None
        }
        """
        if task_id not in self.tasks:
            return {'updated': {}, 'error': f'任务 ID {task_id} 不存在'}

        task = self.tasks[task_id]
        old_start = task.start
        old_planned = task.planned_finish

        # 设置新开始日期
        task.start = new_start
        # 自动计算预计完成日期
        task.planned_finish = calc_planned_finish(new_start, task.duration)
        # 自动推导状态
        task.status = task.derive_status()

        updated = {
            task_id: {
                'old_start': old_start.isoformat() if old_start else None,
                'new_start': task.start.isoformat(),
                'old_planned_finish': old_planned.isoformat() if old_planned else None,
                'new_planned_finish': task.planned_finish.isoformat(),
                'old_status': task.status,
                'new_status': task.status,
            }
        }

        # 获取所有后代并级联更新
        descendants = self.graph.get_descendants(task_id)
        if descendants:
            descendant_order = self.graph.topological_from(descendants)
            for dtid in descendant_order:
                dtask = self.tasks.get(dtid)
                if dtask is None:
                    continue  # 跳过已被删除的任务
                if dtask.duration == 0 and not dtask.task_name:
                    continue  # 跳过纯占位节点

                old_s = dtask.start
                old_pf = dtask.planned_finish
                old_st = dtask.status

                # 尝试从依赖推导开始日期
                computed_start = self.compute_start_date(dtask)
                if computed_start is not None:
                    dtask.start = computed_start
                    dtask.planned_finish = calc_planned_finish(computed_start, dtask.duration)
                    dtask.status = dtask.derive_status()

                    if old_s != dtask.start or old_pf != dtask.planned_finish:
                        updated[dtid] = {
                            'old_start': old_s.isoformat() if old_s else None,
                            'new_start': dtask.start.isoformat(),
                            'old_planned_finish': old_pf.isoformat() if old_pf else None,
                            'new_planned_finish': dtask.planned_finish.isoformat(),
                            'old_status': old_st,
                            'new_status': dtask.status,
                        }

        return {'updated': updated, 'error': None}

    def set_actual_finish(self, task_id: int, actual_finish_str: str | None) -> dict:
        """
        设置/清除实际完成日期。
        有值 → 状态自动变 COMPLETED
        清除 → 状态恢复为 PLANNING（除非是 IN PROGRESS）
        """
        if task_id not in self.tasks:
            return {'updated': {}, 'error': f'任务 ID {task_id} 不存在'}

        task = self.tasks[task_id]
        old_actual = task.actual_finish
        old_status = task.status

        if actual_finish_str:
            task.actual_finish = datetime.datetime.strptime(actual_finish_str, '%Y-%m-%d').date()
        else:
            task.actual_finish = None

        task.status = task.derive_status()

        return {
            'updated': {
                task_id: {
                    'old_actual_finish': old_actual.isoformat() if old_actual else None,
                    'new_actual_finish': task.actual_finish.isoformat() if task.actual_finish else None,
                    'old_status': old_status,
                    'new_status': task.status,
                }
            },
            'error': None
        }

    def update_task_deps(self, task_id: int, pred_ids_str: str, dep_specs_str: str) -> dict:
        """
        ★ 编辑任务的依赖关系。
        
        pred_ids_str: 逗号分隔的纯数字 ID，如 "1,5,6" 或 "" (无前置)
        dep_specs_str: 逗号分隔的依赖规格，如 "FS+3,SS-1" 或 "FS,SS"
                       每个规格格式: <FS|SS>[+-]<天数>，天数部分可选，默认为 0
        
        修改后自动重建依赖图并重算日期。
        """
        if task_id not in self.tasks:
            return {'updated': {}, 'error': f'任务 ID {task_id} 不存在'}

        task = self.tasks[task_id]

        # 解析前置任务 ID
        new_pred_ids = []
        if pred_ids_str.strip():
            for p in pred_ids_str.replace('，', ',').split(','):
                p = p.strip()
                if p:
                    try:
                        pid = int(p)
                        if pid in self.tasks:
                            new_pred_ids.append(pid)
                    except ValueError:
                        return {'updated': {}, 'error': f'无效的前置任务 ID: {p}'}

        # 解析依赖规格 (如 "FS+3", "SS-1", "FS")
        new_pred_types = []
        new_lag_days = []
        specs = [s.strip() for s in dep_specs_str.replace('，', ',').split(',')] if dep_specs_str.strip() else []
        while len(specs) < len(new_pred_ids):
            specs.append('FS')
        specs = specs[:len(new_pred_ids)]

        for spec in specs:
            if not spec:
                new_pred_types.append('FS')
                new_lag_days.append(0)
                continue
            spec_upper = spec.upper().strip()
            if spec_upper.startswith('FS'):
                ptype = 'FS'
                rest = spec[2:].strip()
            elif spec_upper.startswith('SS'):
                ptype = 'SS'
                rest = spec[2:].strip()
            else:
                return {'updated': {}, 'error': f'无效的依赖类型: {spec}，应为 FS 或 SS 开头'}

            if rest:
                try:
                    lag = int(rest)
                except ValueError:
                    return {'updated': {}, 'error': f'无效的延迟天数: {rest}'}
            else:
                lag = 0

            new_pred_types.append(ptype)
            new_lag_days.append(lag)

        old_pred_ids = list(task.pred_task_ids)
        old_pred_types = list(task.pred_types)
        old_lag_days = list(task.lag_days)
        old_start = task.start
        old_planned = task.planned_finish

        task.pred_task_ids = new_pred_ids
        task.pred_types = new_pred_types
        task.lag_days = new_lag_days

        # 重建依赖图
        self.graph._build()

        # 循环检测
        cycle = self.graph.detect_cycle()
        if cycle:
            task.pred_task_ids = old_pred_ids
            task.pred_types = old_pred_types
            task.lag_days = old_lag_days
            self.graph._build()
            task_names = [f"ID {tid}" for tid in cycle]
            return {'updated': {}, 'error': f'检测到循环依赖: {" → ".join(task_names)}，修改已回滚'}

        updated = {}

        if task.start is not None:
            task.planned_finish = calc_planned_finish(task.start, task.duration)
        else:
            computed = self.compute_start_date(task)
            if computed is not None:
                task.start = computed
                task.planned_finish = calc_planned_finish(computed, task.duration)

        task.status = task.derive_status()

        updated[task_id] = {
            'old_pred_ids': old_pred_ids,
            'new_pred_ids': new_pred_ids,
            'old_pred_types': old_pred_types,
            'new_pred_types': new_pred_types,
            'old_lag_days': old_lag_days,
            'new_lag_days': new_lag_days,
            'old_start': old_start.isoformat() if old_start else None,
            'new_start': task.start.isoformat() if task.start else None,
            'old_planned_finish': old_planned.isoformat() if old_planned else None,
            'new_planned_finish': task.planned_finish.isoformat() if task.planned_finish else None,
        }

        descendants = self.graph.get_descendants(task_id)
        if descendants:
            descendant_order = self.graph.topological_from(descendants)
            for dtid in descendant_order:
                dtask = self.tasks.get(dtid)
                if dtask is None:
                    continue
                if dtask.duration == 0 and not dtask.task_name:
                    continue
                old_s = dtask.start
                old_pf = dtask.planned_finish
                computed_start = self.compute_start_date(dtask)
                if computed_start is not None:
                    dtask.start = computed_start
                    dtask.planned_finish = calc_planned_finish(computed_start, dtask.duration)
                    dtask.status = dtask.derive_status()
                    if old_s != dtask.start or old_pf != dtask.planned_finish:
                        updated[dtid] = {
                            'old_start': old_s.isoformat() if old_s else None,
                            'new_start': dtask.start.isoformat(),
                            'old_planned_finish': old_pf.isoformat() if old_pf else None,
                            'new_planned_finish': dtask.planned_finish.isoformat(),
                        }

        return {'updated': updated, 'error': None}

    def set_status(self, task_id: int, new_status: str) -> dict:
        """
        手动设置任务状态。
        允许 PLANNING / IN PROGRESS / COMPLETED 三向切换，
        即使有实际完成日期也允许手动修改状态。
        """
        if task_id not in self.tasks:
            return {'updated': {}, 'error': f'任务 ID {task_id} 不存在'}

        task = self.tasks[task_id]
        old_status = task.status

        valid_statuses = {'PLANNING', 'IN PROGRESS', 'COMPLETED'}
        if new_status not in valid_statuses:
            return {'updated': {}, 'error': f'无效状态: {new_status}，有效值: {valid_statuses}'}

        task.status = new_status

        return {
            'updated': {
                task_id: {
                    'old_status': old_status,
                    'new_status': new_status,
                }
            },
            'error': None
        }

    def delete_task(self, task_id: int) -> dict:
        """
        删除任务。自动处理：
        1. 将该任务的后继任务的前置依赖重新指向该任务的前置任务
        2. 从依赖图中移除该任务
        3. 重算受影响任务的日期
        """
        if task_id not in self.tasks:
            return {'updated': {}, 'error': f'任务 ID {task_id} 不存在'}

        task = self.tasks[task_id]
        deleted_name = task.task_name or f'(辅助节点)'

        # 收集受影响的后续任务（在删除前）
        affected_descendants = self.graph.get_descendants(task_id)

        # 处理该任务的后继任务：将它们对该任务的依赖替换为该任务的前置依赖
        successors = list(self.graph.successors.get(task_id, []))
        for succ_id in successors:
            succ = self.tasks[succ_id]
            # 找到 succ 中指向被删任务的依赖索引
            new_pred_ids = []
            new_pred_types = []
            new_lag_days = []
            for i, pid in enumerate(succ.pred_task_ids):
                if pid == task_id:
                    # 用被删任务的前置依赖替换
                    for j, grand_pid in enumerate(task.pred_task_ids):
                        if grand_pid not in new_pred_ids:  # 避免重复
                            new_pred_ids.append(grand_pid)
                            new_pred_types.append(task.pred_types[j] if j < len(task.pred_types) else 'FS')
                            new_lag_days.append(task.lag_days[j] if j < len(task.lag_days) else 0)
                else:
                    if pid not in new_pred_ids:
                        new_pred_ids.append(pid)
                        new_pred_types.append(succ.pred_types[i] if i < len(succ.pred_types) else 'FS')
                        new_lag_days.append(succ.lag_days[i] if i < len(succ.lag_days) else 0)
            succ.pred_task_ids = new_pred_ids
            succ.pred_types = new_pred_types
            succ.lag_days = new_lag_days

        # 从 tasks 中移除
        del self.tasks[task_id]

        # 重建依赖图
        self.graph._build()

        # 重算受影响任务
        updated = {task_id: {'action': 'deleted', 'task_name': deleted_name}}
        all_affected = affected_descendants | set(successors)
        all_affected.discard(task_id)

        if all_affected:
            order = self.graph.topological_from(all_affected)
            for dtid in order:
                dtask = self.tasks.get(dtid)
                if not dtask or (dtask.duration == 0 and not dtask.task_name):
                    continue
                old_s = dtask.start
                old_pf = dtask.planned_finish
                computed = self.compute_start_date(dtask)
                if computed is not None:
                    dtask.start = computed
                    dtask.planned_finish = calc_planned_finish(computed, dtask.duration)
                    dtask.status = dtask.derive_status()
                    if old_s != dtask.start or old_pf != dtask.planned_finish:
                        updated[dtid] = {
                            'old_start': old_s.isoformat() if old_s else None,
                            'new_start': dtask.start.isoformat(),
                            'old_planned_finish': old_pf.isoformat() if old_pf else None,
                            'new_planned_finish': dtask.planned_finish.isoformat(),
                        }

        return {'updated': updated, 'error': None}


# ═══════════════════════════════════════════════════════════════
# Excel 输出器 v2
# ═══════════════════════════════════════════════════════════════

class ExcelWriter:
    """
    将更新后的任务数据写回 Excel。
    Schedule 工作表：新增 Actual Finish 列 (J)，更新列标题
    Gantt 工作表：保留原公式（自动引用 Schedule 数据）
    """

    # 新的列布局
    # A: Task ID, B: Task Name, C: Duration, D: Pred Task ID
    # E: Pred Type, F: Lag, G: Start, H: Planned Finish (原 Finish)
    # I: Actual Finish (新增), J: Status

    def __init__(self, template_path: str, tasks: dict[int, Task],
                 project_start: datetime.date, project_address: str):
        self.template_path = template_path
        self.tasks = tasks
        self.project_start = project_start
        self.project_address = project_address

    def write(self, output_path: str):
        """生成输出 Excel 文件"""
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb['Schedule']

        # 更新列标题
        ws['H3'] = 'Planned Finish'         # 原 "Finish" → "Planned Finish"
        ws['I3'] = 'Actual Finish'          # 新增列标题
        ws['J3'] = 'Status'                 # 状态列标题（原本就在 I 列，现在移到 J）

        # 清除原有 G/H/I 列数据，写入新数据
        for tid, task in self.tasks.items():
            row = task.excel_row
            # G: Start (用户输入的计划开始日期)
            ws.cell(row=row, column=7).value = task.start
            # H: Planned Finish (自动计算的预计完成日期)
            ws.cell(row=row, column=8).value = task.planned_finish
            # I: Actual Finish (用户输入的实际完成日期)
            ws.cell(row=row, column=9).value = task.actual_finish
            # J: Status
            ws.cell(row=row, column=10).value = task.status

        # 清除第 4 行到末尾的旧状态数据（J 列之外也清一下 I 列残留）
        # Gantt 工作表保持不动，公式会自动引用 Schedule 的新日期

        wb.save(output_path)
        print(f"✅ 已保存到: {output_path}")
        return output_path


# ═══════════════════════════════════════════════════════════════
# 项目管理器（顶层接口）v2
# ═══════════════════════════════════════════════════════════════

class ProjectManager:
    """项目管理器 - 顶层接口，支持数据持久化"""

    def __init__(self, excel_path: str, data_dir: str | None = None):
        self.excel_path = excel_path
        self.data_dir = data_dir or os.path.dirname(excel_path)
        self.snapshot_path = os.path.join(self.data_dir, '.schedule_snapshot.json')
        self.tasks: dict[int, Task] = {}
        self.graph: DependencyGraph | None = None
        self.engine: ScheduleEngine | None = None
        self.project_start: datetime.date | None = None
        self.project_address: str = ""

    def load(self):
        """加载数据：优先从快照恢复，否则从 Excel 解析（首次清空日期）"""
        if os.path.exists(self.snapshot_path):
            try:
                self._load_snapshot()
                self.graph = DependencyGraph(self.tasks)
                self.engine = ScheduleEngine(self.tasks, self.graph, self.project_start)
                return self
            except Exception as e:
                print(f"快照加载失败，从 Excel 重新加载: {e}")

        # 首次加载：从 Excel 解析，清空日期
        parser = ScheduleParser(self.excel_path).parse()
        self.tasks = parser.tasks
        self.project_start = parser.project_start_date
        self.project_address = parser.project_address

        self.graph = DependencyGraph(self.tasks)
        cycle = self.graph.detect_cycle()
        if cycle:
            task_names = [f"#{tid} ({self.tasks[tid].task_name})" for tid in cycle]
            raise ValueError(f"检测到循环依赖: {' → '.join(task_names)}")

        self.engine = ScheduleEngine(self.tasks, self.graph, self.project_start)

        # 自动删除辅助任务 77, 78
        for auto_del_id in [77, 78]:
            if auto_del_id in self.tasks:
                self.engine.delete_task(auto_del_id)

        # 首次加载后保存快照
        self._save_snapshot()
        return self

    def _task_to_dict(self, task: Task) -> dict:
        return {
            'task_id': task.task_id,
            'task_name': task.task_name,
            'duration': task.duration,
            'pred_task_ids': task.pred_task_ids,
            'pred_types': task.pred_types,
            'lag_days': task.lag_days,
            'start': task.start.isoformat() if task.start else None,
            'planned_finish': task.planned_finish.isoformat() if task.planned_finish else None,
            'actual_finish': task.actual_finish.isoformat() if task.actual_finish else None,
            'status': task.status,
            'excel_row': task.excel_row,
        }

    def _dict_to_task(self, d: dict) -> Task:
        return Task(
            task_id=d['task_id'],
            task_name=d['task_name'],
            duration=d['duration'],
            pred_task_ids=d['pred_task_ids'],
            pred_types=d['pred_types'],
            lag_days=d['lag_days'],
            start=datetime.date.fromisoformat(d['start']) if d['start'] else None,
            planned_finish=datetime.date.fromisoformat(d['planned_finish']) if d['planned_finish'] else None,
            actual_finish=datetime.date.fromisoformat(d['actual_finish']) if d['actual_finish'] else None,
            status=d['status'],
            excel_row=d['excel_row'],
        )

    def _save_snapshot(self):
        """保存当前状态到 JSON 快照"""
        data = {
            'project_start': self.project_start.isoformat() if self.project_start else None,
            'project_address': self.project_address,
            'tasks': [self._task_to_dict(t) for t in self.tasks.values()],
        }
        with open(self.snapshot_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _load_snapshot(self):
        """从 JSON 快照恢复状态"""
        with open(self.snapshot_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        self.project_start = datetime.date.fromisoformat(data['project_start']) if data['project_start'] else None
        self.project_address = data.get('project_address', '')
        self.tasks = {}
        for td in data['tasks']:
            task = self._dict_to_task(td)
            self.tasks[task.task_id] = task

    def update_start(self, task_id: int, new_start_str: str) -> dict:
        """用户输入开始日期 → 自动计算预计完成并级联更新"""
        new_start = datetime.datetime.strptime(new_start_str, '%Y-%m-%d').date()
        result = self.engine.update_task_start(task_id, new_start)
        self._save_snapshot()
        return result

    def set_actual_finish(self, task_id: int, actual_finish_str: str | None) -> dict:
        """设置或清除实际完成日期"""
        result = self.engine.set_actual_finish(task_id, actual_finish_str)
        self._save_snapshot()
        return result

    def set_status(self, task_id: int, new_status: str) -> dict:
        """手动修改状态"""
        result = self.engine.set_status(task_id, new_status)
        self._save_snapshot()
        return result

    def update_deps(self, task_id: int, pred_ids_str: str, dep_specs_str: str) -> dict:
        """编辑前置任务和依赖类型"""
        result = self.engine.update_task_deps(task_id, pred_ids_str, dep_specs_str)
        self._save_snapshot()
        return result

    def delete_task(self, task_id: int) -> dict:
        """删除任务"""
        result = self.engine.delete_task(task_id)
        self._save_snapshot()
        return result

    def reset(self):
        """重置：删除快照，重新从 Excel 加载"""
        if os.path.exists(self.snapshot_path):
            os.remove(self.snapshot_path)
        return self.load()

    def save(self, output_path: str):
        """导出 Excel（使用当前最新数据）"""
        writer = ExcelWriter(
            self.excel_path, self.tasks,
            self.project_start, self.project_address
        )
        writer.write(output_path)

    def get_task_list(self) -> list[dict]:
        """获取任务列表（供前端）"""
        result = []
        for tid in sorted(self.tasks.keys()):
            task = self.tasks[tid]
            if not task.task_name and task.duration == 0:
                continue
            result.append({
                'task_id': task.task_id,
                'task_name': task.task_name,
                'duration': task.duration,
                'pred_ids': task.pred_task_ids,
                'pred_types': task.pred_types,
                'lag_days': task.lag_days,
                'dep_specs': [f"{task.pred_types[i]}{task.lag_days[i]:+d}" if task.lag_days[i] else task.pred_types[i]
                              for i in range(len(task.pred_task_ids))],
                'dep_specs_str': ','.join(
                    f"{task.pred_types[i]}{task.lag_days[i]:+d}" if task.lag_days[i] else task.pred_types[i]
                    for i in range(len(task.pred_task_ids))
                ),
                'pred_ids_str': ','.join(str(pid) for pid in task.pred_task_ids),
                'start': task.start.isoformat() if task.start else None,
                'planned_finish': task.planned_finish.isoformat() if task.planned_finish else None,
                'actual_finish': task.actual_finish.isoformat() if task.actual_finish else None,
                'status': task.status,
                'excel_row': task.excel_row,
            })
        return result

    def get_summary(self) -> dict:
        """获取项目摘要"""
        valid = [t for t in self.tasks.values() if t.start and t.planned_finish and t.task_name]
        completed = [t for t in self.tasks.values() if t.status == 'COMPLETED' and t.task_name]
        in_progress = [t for t in self.tasks.values() if t.status == 'IN PROGRESS' and t.task_name]
        if not valid:
            return {
                'project_address': self.project_address,
                'project_start': self.project_start.isoformat() if self.project_start else None,
                'total_tasks': len([t for t in self.tasks.values() if t.task_name]),
                'earliest_start': None,
                'latest_finish': None,
                'completed_count': len(completed),
                'in_progress_count': len(in_progress),
            }
        starts = [t.start for t in valid]
        finishes = [t.planned_finish for t in valid]
        return {
            'project_address': self.project_address,
            'project_start': self.project_start.isoformat() if self.project_start else None,
            'total_tasks': len([t for t in self.tasks.values() if t.task_name]),
            'earliest_start': min(starts).isoformat(),
            'latest_finish': max(finishes).isoformat(),
            'completed_count': len(completed),
            'in_progress_count': len(in_progress),
        }
